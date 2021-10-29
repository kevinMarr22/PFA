import openpyxl
import smtplib
import mimetypes
import ssl
import os
from email.message import EmailMessage
def pedirNumeroEntero():
 
    correcto=False
    num=0
    while(not correcto):
        try:
            num = int(input("Introduce un numero entero: "))
            correcto=True
        except ValueError:
            print('Error, introduce un numero entero')
     
    return num
 
salir = False
opcion = 0
 
while not salir:
 
    print("a que hoja desea ingresar")
    print ("1. clientes")
    print ("2. productos")
    print ("3. pedidos")
    print ("4. copia de seguridad")
    opcion = pedirNumeroEntero()
 
    if opcion == 1:
        menu = """
        Bienvenidos al registro de productos, llene los campos que usted
        prefiera a continuacion seleccionando un numero del 1 al 3:
        [1] Añadir Cliente 
        [2] Listar clientes
        [3] enviar cotizacion por correo electronico  
        """ 
        print(menu)
        opcion = input('Digita una opcion entre 1 y 3: ')
        if opcion == '1':
            libro = openpyxl.load_workbook('ejemplo.xlsx')
            hoja = libro['clientes']

            hoja['A1'].value='Nombre'
            hoja['B1'].value='nit'
            hoja['C1'].value='direccion'

            consultar=""
            clientes=[]
            while consultar !="N":
             nueva_cliente=input("ingrese nombre")
             clientes.append(nueva_cliente)

             clientes1=[]
             nueva_cliente1=input("ingrese nit")
             clientes1.append(nueva_cliente1)
             
             clientes2=[]
             nueva_cliente2=input("ingrese direccion")
             clientes2.append(nueva_cliente2)
             consultar=input("precione. N")

            data={"Nombre":clientes,
                  "nit":clientes1,
                  "direccion":clientes2
                  }
            print(data)

            primeraFila=hoja.max_row + 1
            for cliente in clientes:
             hoja['A' + str(primeraFila)].value=cliente
            for cliente in clientes1:
             hoja['B' + str(primeraFila)].value=cliente
            for cliente in clientes2:
             hoja['C' + str(primeraFila)].value=cliente
             primeraFila=primeraFila+1

            libro.save('ejemplo.xlsx')
            libro.save('inventario.xlsx')
        elif opcion == '2':
            libro = openpyxl.load_workbook("ejemplo.xlsx")
            hojaClientes=libro['clientes']
            diccionarioClientes={}
            clientes=[]
            for row in range(2, hojaClientes.max_row +1): 
                diccionarioClientes['Nombre']=hojaClientes["A" + str(row)].value
                diccionarioClientes['nit']=hojaClientes["B" + str(row)].value
                diccionarioClientes['direccion']=hojaClientes["C" + str(row)].value
                clientes.append(diccionarioClientes)
                diccionarioClientes={}
            for item in clientes:
                print("Nombre: " + str(item['Nombre']))
                print("nit: " + str(item['nit']))
                print("direccion: " + str(item['direccion']))
        elif opcion == '3':
            Direccion_de_servidor = "smtp.gmail.com"
            puerto = "587"
            Direccion_de_origen = "elrisas.jsss@gmail.com" #gmail desde donde se envia el correo
            contraseña = "el123risas" #contraseña del gmail lo podes usar men

            C=''
            C=input('ingrese el correo')
            mensaje = EmailMessage()
            mensaje["Subject"] = "Hola como estas "
            mensaje ["From"] = Direccion_de_origen
            mensaje["To"] = C #aqui podes poner una variable que pregunte el correo al cual enviar el archivo y lo remplazas por el correo  

            mensaje.set_content("Cuerpo del mensaje")

            mensaje.add_alternative("""     
            <p> 
                <h1>Copia de respaldo </h1>
                Hola <strong> archivo adjunto "proyecto" </strong>
            </p>

            """, subtype = "html")

            Nombre_archivo = "cotizacion.xlsx" #Modificar por el archivo de exel donde esta el listado "Base_de_datos.xlsx" o como le tengas xd

            #no modificar 
            ctype, encoding = mimetypes.guess_type(Nombre_archivo) #tuplas
            if ctype is None or encoding is not None:
                ctype = 'application/octet-stream'

            Tipo_principal, subTipo = ctype.split('/',1)
            with open(Nombre_archivo, 'rb') as archivo_leido:
                mensaje.add_attachment(archivo_leido.read(), maintype=Tipo_principal, subtype = subTipo, filename = Nombre_archivo)

            context = ssl.create_default_context()

            smtp = smtplib.SMTP(Direccion_de_servidor, puerto)
            smtp.starttls()
            smtp.login(Direccion_de_origen, contraseña)
            smtp.send_message(mensaje)
        else:
            print('Debes digitar un numero entre 1 y 3')
            print('=-='*20)
    elif opcion == 2:
        menu = """
        Bienvenidos al registro de productos, llene los campos que usted
        prefiera a continuacion seleccionando un numero del 1 al 3:
        [1] Añadir producto 
        [2] Listar productos
        """ 
        print(menu)
        opcion = input('Digita una opcion entre 1 y 3: ')
        if opcion == '1':
            libro = openpyxl.load_workbook('ejemplo.xlsx')
            hoja = libro['productos']

            hoja['A1'].value='Producto'
            hoja['B1'].value='Precio'
            hoja['C1'].value='Existencia'

            consultar=""
            clientesp=[]
            while consultar !="N":
             nueva_clientep=input("ingrese producto")
             clientesp.append(nueva_clientep)

             clientesp1=[]
             nueva_clientep1=input("ingrese precio")
             clientesp1.append(nueva_clientep1)

             clientesp2=[] 
             nueva_clientep2=input("ingrese existencia")
             clientesp2.append(nueva_clientep2)

             consultar=input("precione. N")

            datap={"Producto":clientesp,
                "Precio":clientesp1,
                "Existencia":clientesp2
            }

            print(datap)

            primeraFila=hoja.max_row + 1
            for clientep in clientesp:
             hoja['A' + str(primeraFila)].value=clientep
            for clientep in clientesp1:
             hoja['B' + str(primeraFila)].value=clientep
            for clientep in clientesp2:
             hoja['C' + str(primeraFila)].value=clientep
            primeraFila=primeraFila+1

            libro.save('ejemplo.xlsx')
            libro.save('inventario.xlsx')
            libro.save('cotizacion.xlsx')

        elif opcion == '2':
            libro = openpyxl.load_workbook("ejemplo.xlsx")
            hojaProductos=libro['productos']
            diccionarioProductos={}
            Productos=[]
            for row in range(2, hojaProductos.max_row +1): 
                diccionarioProductos['productos']=hojaProductos["A" + str(row)].value
                diccionarioProductos['Precio']=hojaProductos["B" + str(row)].value
                diccionarioProductos['Existencia']=hojaProductos["C" + str(row)].value
                Productos.append(diccionarioProductos)
                diccionarioProductos={}
            for item in Productos:
                print("productos: " + str(item['productos']))
                print("precio: " + str(item['Precio']))
                print("existencia: " + str(item['Existencia']))
        else:
            print('Debes digitar un numero entre 1 y 3')
            print('=-='*20)
    elif opcion == 3:
        menu = """
        Bienvenidos al registro de productos, llene los campos que usted
        prefiera a continuacion seleccionando un numero del 1 al 3:
        [1] Añadir producto 
        [2] Listar productos 
        """ 
        print(menu)
        opcion = input('Digita una opcion entre 1 y 3: ')
        if opcion == '1':
            libro = openpyxl.load_workbook('ejemplo.xlsx')
            hoja = libro['ventas']
            hoja['A1'].value='cliente'
            hoja['B1'].value='producto'
            hoja['C1'].value='cantidad'
            hoja['D1'].value='valor'

            consultar=""
            clientesv=[]
            while consultar !="N":
             nueva_clientev=input("ingrese cliente")
             clientesv.append(nueva_clientev)

             clientesv1=[]
             nueva_clientev1=input("ingrese producto")
             clientesv1.append(nueva_clientev1)

             clientesv2=[] 
             nueva_clientev2=input("ingrese cantidad")
             clientesv2.append(nueva_clientev2)

             clientesv3=[] 
             nueva_clientev3=input("ingrese precio")
             clientesv3.append(nueva_clientev3)
             consultar=input("precione. N")

            datap={"cliente":clientesv,
                "producto":clientesv1,
                "cantidad":clientesv2,
                "precio":clientesv3
            }

            print(datap)

            primeraFila=hoja.max_row + 1
            for clientev in clientesv:
                hoja['A' + str(primeraFila)].value=clientev
            for clientev in clientesv1:
                hoja['B' + str(primeraFila)].value=clientev
            for clientev in clientesv2:
                hoja['C' + str(primeraFila)].value=clientev
            for clientev in clientesv3:
                hoja['D' + str(primeraFila)].value=clientev
            primeraFila=primeraFila+1

            libro.save('ejemplo.xlsx')
            libro.save('inventario.xlsx')

        elif opcion == '2':
            libro = openpyxl.load_workbook("ejemplo.xlsx")
            hojaVentas=libro['ventas']
            diccionarioVentas={}
            Ventas=[]
            for row in range(2, hojaVentas.max_row +1): 
                diccionarioVentas['cliente']=hojaVentas["A" + str(row)].value
                diccionarioVentas['producto']=hojaVentas["B" + str(row)].value
                diccionarioVentas['precio']=hojaVentas["C" + str(row)].value
                Ventas.append(diccionarioVentas)
                diccionarioVentas={}
            for item in Ventas:
                print("cliente: " + str(item['cliente']))
                print("producto: " + str(item['producto']))
                print("precio: " + str(item['precio']))
        else:
            print('Debes digitar un numero entre 1 y 3')
            print('=-='*20)   
    elif opcion == 4:
        Direccion_de_servidor = "smtp.gmail.com"
        puerto = "587"
        Direccion_de_origen = "elrisas.jsss@gmail.com" #gmail desde donde se envia el correo
        contraseña = "el123risas" #contraseña del gmail lo podes usar men

        mensaje = EmailMessage()
        mensaje["Subject"] = "Hola como estas "
        mensaje ["From"] = Direccion_de_origen
        mensaje["To"] = "kmarroquing4@miumg.edu.gt"#aqui podes poner una variable que pregunte el correo al cual enviar el archivo y lo remplazas por el correo  

        mensaje.set_content("Cuerpo del mensaje")

        mensaje.add_alternative("""     
        <p> 
            <h1>Copia de respaldo </h1>
            Hola <strong> archivo adjunto "proyecto" </strong>
        </p>

        """, subtype = "html")

        Nombre_archivo = "inventario.xlsx" #Modificar por el archivo de exel donde esta el listado "Base_de_datos.xlsx" o como le tengas xd

        #no modificar nada apartir de aqui
        ctype, encoding = mimetypes.guess_type(Nombre_archivo) #tuplas
        if ctype is None or encoding is not None:
            ctype = 'application/octet-stream'

        Tipo_principal, subTipo = ctype.split('/',1)
        with open(Nombre_archivo, 'rb') as archivo_leido:
            mensaje.add_attachment(archivo_leido.read(), maintype=Tipo_principal, subtype = subTipo, filename = Nombre_archivo)

        context = ssl.create_default_context()

        smtp = smtplib.SMTP(Direccion_de_servidor, puerto)
        smtp.starttls()
        smtp.login(Direccion_de_origen, contraseña)
        smtp.send_message(mensaje)
    else:
        print ("Introduce un numero entre 1 y 3")
print ("Fin")